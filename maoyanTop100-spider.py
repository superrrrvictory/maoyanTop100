from requests_html import HTMLSession
import openpyxl

session = HTMLSession()
headers = {'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
           'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.92 Safari/537.36'}

# 解析页面
def parse_one_page(web_url):
    # 获取标题
    r = session.get(web_url, headers=headers)
    titles = []
    targets = r.html.xpath('//*[@id="app"]/div/div/div[1]/dl/dd/div/div/div[1]/p[1]/a')
    for target in targets:
        titles.append(target.text)

    # 获取演员信息
    actors = []
    targets1 = r.html.xpath('//*[@id="app"]/div/div/div[1]/dl/dd/div/div/div[1]/p[2]')
    for target1 in targets1:
        actors.append(target1.text)

    # 获取上映时间
    times = []
    targets2 = r.html.xpath('//*[@id="app"]/div/div/div[1]/dl/dd/div/div/div[1]/p[3]')
    for target2 in targets2:
        times.append(target2.text)

    # 获取评分
    scores = []
    targets3 = r.html.xpath('//*[@id="app"]/div/div/div[1]/dl/dd/div/div/div[2]/p')
    for target3 in targets3:
        scores.append(target3.text)

        # result = zip([titles, actors, times, scores])

    # 输出结果
    results = []
    length = len(titles)
    for j in range(length):
        results.append([titles[j], scores[j], actors[j], times[j]])

    return results


# 保存到文档中
def save_to_my_computer(result):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = '电影名称'

    ws['B1'] = '评分'

    ws['C1'] = '主演'

    ws['D1'] = '上映时间'

    for item in result:
        ws.append(item)

    wb.save(u'猫眼电影TOP100.xlsx')


# 定义主函数
def main():
    result = []
    urls = []
    for i in range(10):
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.116 Safari/537.36'}
        urls.append('http://maoyan.com/board/4?offset=' + str(i * 10))
    for url in urls:
        html = parse_one_page(url)
        result.extend(html)
    print(result)
    print(len(result))
    print(urls)
    save_to_my_computer(result)

if __name__ == "__main__":
    main()
